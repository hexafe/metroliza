"""Correctness boundaries for request-local histogram reuse and artifact proof."""
from __future__ import annotations

from pathlib import Path
import shutil
import zipfile

import numpy as np
import openpyxl
import pytest

from metroliza.charts import hexafe_plotstats_adapter as adapter
from scripts.compare_csv_pipeline_artifacts import (
    _normalized_part, artifact_manifest, compare_artifacts,
)


def _count_table_computations(monkeypatch):
    calls = []

    def compute(values, *, lsl=None, usl=None):
        calls.append((tuple(values), lsl, usl))
        return (("Mean", str(float(np.mean(values)))), ("Limits", repr((lsl, usl))))

    monkeypatch.setattr(adapter, "_histogram_table_rows_from_plotstats", compute)
    return calls


def test_request_cache_uses_input_contents_limits_and_separate_titles(monkeypatch):
    calls = _count_table_computations(monkeypatch)
    values = np.array([1.0, 2.0, 3.0])
    with adapter.histogram_stats_request():
        first = adapter.build_histogram_stats_table(values, title="First", lsl=0.0, usl=4.0)
        same = adapter.build_histogram_stats_table(values.copy(), title="Second", lsl=0.0, usl=4.0)
        assert first.rows == same.rows
        assert (first.title, same.title) == ("First", "Second")
        assert len(calls) == 1
        values[0] = 10.0
        changed = adapter.build_histogram_stats_table(values, lsl=0.0, usl=4.0)
        assert changed.rows != first.rows
        reordered = adapter.build_histogram_stats_table(values[::-1], lsl=0.0, usl=4.0)
        assert reordered.rows == changed.rows
        lower = adapter.build_histogram_stats_table(values, lsl=1.0, usl=4.0)
        upper = adapter.build_histogram_stats_table(values, lsl=1.0, usl=5.0)
        assert lower.rows != upper.rows
        assert len(calls) == 5
    with adapter.histogram_stats_request():
        adapter.build_histogram_stats_table(values, lsl=1.0, usl=5.0)
    assert len(calls) == 6
    assert adapter._histogram_table_cache.get() is None


def test_request_cache_preserves_coercion_missing_values_and_empty_input(monkeypatch):
    calls = _count_table_computations(monkeypatch)
    with adapter.histogram_stats_request():
        mixed = adapter.build_histogram_stats_table(["1", None, "bad", np.inf, 2.0, np.nan])
        numeric = adapter.build_histogram_stats_table([1.0, 2.0])
        assert mixed.rows == numeric.rows
        assert adapter.build_histogram_stats_table([None, "bad", np.nan]) is None
        assert len(calls) == 1


def test_request_cache_preserves_signed_zero_in_limit_settings(monkeypatch):
    calls = _count_table_computations(monkeypatch)
    with adapter.histogram_stats_request():
        positive = adapter.build_histogram_stats_table([1.0, 2.0], lsl=0.0, usl=3.0)
        negative = adapter.build_histogram_stats_table([1.0, 2.0], lsl=-0.0, usl=3.0)
        assert positive.rows != negative.rows
        assert len(calls) == 2


@pytest.mark.parametrize("failure", [RuntimeError("render failure"), KeyboardInterrupt()])
def test_request_cache_releases_buffers_and_restores_nested_context(monkeypatch, failure):
    calls = _count_table_computations(monkeypatch)
    with adapter.histogram_stats_request():
        adapter.build_histogram_stats_table([1.0, 2.0])
        outer = adapter._histogram_table_cache.get()
        with pytest.raises(type(failure)):
            with adapter.histogram_stats_request():
                adapter.build_histogram_stats_table([1.0, 2.0])
                inner = adapter._histogram_table_cache.get()
                raise failure
        assert not inner.rows and inner.payload_bytes == 0
        assert adapter._histogram_table_cache.get() is outer
        adapter.build_histogram_stats_table([1.0, 2.0])
        assert len(calls) == 2
    assert not outer.rows and outer.payload_bytes == 0
    assert adapter._histogram_table_cache.get() is None


def test_request_cache_bounds_entries_and_bytes_without_changing_results(monkeypatch):
    calls = _count_table_computations(monkeypatch)
    monkeypatch.setattr(adapter, "_HISTOGRAM_TABLE_CACHE_MAX_ENTRIES", 1)
    with adapter.histogram_stats_request():
        adapter.build_histogram_stats_table([1.0, 2.0])
        expected = adapter.build_histogram_stats_table([3.0, 4.0])
        assert adapter.build_histogram_stats_table([3.0, 4.0]) == expected
        assert len(calls) == 3
        assert len(adapter._histogram_table_cache.get().rows) == 1
    monkeypatch.setattr(adapter, "_HISTOGRAM_TABLE_CACHE_MAX_BYTES", 1)
    with adapter.histogram_stats_request():
        adapter.build_histogram_stats_table([1.0, 2.0])
        adapter.build_histogram_stats_table([1.0, 2.0])
        assert not adapter._histogram_table_cache.get().rows
    assert len(calls) == 5
    # Input fits, but the returned rows exceed the remaining payload budget.
    monkeypatch.setattr(adapter, "_HISTOGRAM_TABLE_CACHE_MAX_BYTES", 16)
    with adapter.histogram_stats_request():
        adapter.build_histogram_stats_table([1.0, 2.0])
        assert not adapter._histogram_table_cache.get().rows


def test_request_cache_does_not_remember_unavailable_backend(monkeypatch):
    monkeypatch.setattr(adapter, "_histogram_table_rows_from_plotstats", lambda *a, **k: ())
    with adapter.histogram_stats_request():
        fallback = adapter.build_histogram_stats_table([1.0, 2.0], backend="python-fallback")
        assert fallback.backend == "python-fallback"
        assert not adapter._histogram_table_cache.get().rows
        calls = _count_table_computations(monkeypatch)
        recovered = adapter.build_histogram_stats_table([1.0, 2.0])
        assert recovered.backend == "hexafe-plotstats"
        assert len(calls) == 1
    monkeypatch.setattr(adapter, "_histogram_table_rows_from_plotstats", lambda *a, **k: ())
    with adapter.histogram_stats_request():
        assert adapter.build_histogram_stats_table([1.0, 2.0], backend="python").backend == "python"


def test_real_histogram_table_is_identical_with_reuse_and_new_requests():
    values = [1.2, 1.4, None, 1.7, "invalid", 2.1, 2.2, 2.6, np.inf]
    for lsl, usl in ((None, None), (1.0, 3.0), (None, 2.0), (-2.0, None)):
        reference = adapter.build_histogram_stats_table(values, lsl=lsl, usl=usl)
        with adapter.histogram_stats_request():
            assert adapter.build_histogram_stats_table(values, lsl=lsl, usl=usl) == reference
            assert adapter.build_histogram_stats_table(list(values), lsl=lsl, usl=usl) == reference


def test_workflow_request_cancellation_clears_cache_before_next_request(tmp_path, monkeypatch):
    from metroliza.industrial.industrial_analytics_workflow import (
        AnalyticsCancelled, run_tabular_file_analytics,
    )

    calls = _count_table_computations(monkeypatch)
    source = tmp_path / "source.csv"
    source.write_text("PART,DIM\nA,1\nB,2\n")
    retained = []

    def cancel():
        adapter.build_histogram_stats_table([1.0, 2.0])
        retained.append(adapter._histogram_table_cache.get())
        return True

    for _ in range(2):
        with pytest.raises(AnalyticsCancelled):
            run_tabular_file_analytics(
                input_file=str(source), output_dashboard_file=str(tmp_path / "dashboard.html"),
                cancel_check=cancel,
            )
        assert adapter._histogram_table_cache.get() is None
        assert not retained[-1].rows and retained[-1].payload_bytes == 0
    assert len(calls) == 2
    assert not (tmp_path / "dashboard.html").exists()


def test_successive_workflows_use_new_data_limits_and_chart_selection(tmp_path):
    from metroliza.industrial.industrial_analytics_state import (
        ProductionChartSelection, ProductionMetricSelection,
    )
    from metroliza.industrial.industrial_analytics_workflow import run_tabular_file_analytics

    source = tmp_path / "source.csv"
    for index, offset in enumerate((0, 10)):
        source.write_text("PART,DIM\n" + "".join(f"P{i},{i + offset}\n" for i in range(1, 5)))
        output = tmp_path / str(index)
        output.mkdir()
        result = run_tabular_file_analytics(
            input_file=str(source), output_dashboard_file=str(output / "dashboard.html"),
            output_workbook_file=str(output / "workbook.xlsx"), reference_column="PART",
            metric_selection=(ProductionMetricSelection(
                "dim", display_label=f"Metric <{index}>", lsl=float(offset), usl=float(offset + 5),
            ),),
            chart_selection=ProductionChartSelection(
                time_series=False, histogram=True, violin=False, box=bool(index), groupstats=False,
            ),
        )
        assert result.row_count == 4
        assert result.html_dashboard_chart_count == index + 1
        assert adapter._histogram_table_cache.get() is None
        html = (output / "dashboard.html").read_text()
        assert f"Metric &lt;{index}&gt;" in html
        assert f"<td>Mean</td><td>{2.5 + offset:.3f}</td>" in html
        workbook = openpyxl.load_workbook(output / "workbook.xlsx", read_only=True)
        try:
            rows = workbook["Table Data"].iter_rows(values_only=True)
            headers = next(rows)
            assert [row[headers.index("dim")] for row in rows] == [offset + i for i in range(1, 5)]
            rows = workbook["Metrics"].iter_rows(values_only=True)
            headers = next(rows)
            assert next(rows)[headers.index("mean")] == offset + 2.5
        finally:
            workbook.close()


def _minimal_artifacts(directory: Path) -> None:
    import xlsxwriter

    directory.mkdir()
    (directory / "runtime.js").write_text("/* synthetic offline asset */")
    (directory / "dashboard.html").write_text('<script src="runtime.js"></script><p>&lt;safe&gt;</p>')
    with xlsxwriter.Workbook(directory / "workbook.xlsx", {"strings_to_formulas": False}) as workbook:
        sheet = workbook.add_worksheet("Table Data")
        sheet.write_row(0, 0, ["Reference", "Value"])
        sheet.write_row(1, 0, ["=1+1", 1.5])
        sheet.write_row(2, 0, ["/tmp/metroliza_csv_summary_original.sqlite", 2.5])
        chart = workbook.add_chart({"type": "column"})
        chart.add_series({"values": "='Table Data'!$B$2:$B$3"})
        sheet.insert_chart("D1", chart)


@pytest.mark.parametrize("mutation", ["cell", "formula", "chart", "asset", "html", "path_text"])
def test_artifact_comparison_rejects_meaningful_output_loss(tmp_path, mutation):
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _minimal_artifacts(baseline)
    shutil.copytree(baseline, candidate)
    assert compare_artifacts(baseline, candidate)["equal"]
    if mutation in {"cell", "formula", "path_text"}:
        workbook = openpyxl.load_workbook(candidate / "workbook.xlsx")
        if mutation == "path_text":
            workbook.active["A3"] = "/tmp/metroliza_csv_summary_changed.sqlite"
        else:
            workbook.active["B2"] = 9.0 if mutation == "cell" else "=1+1"
        workbook.save(candidate / "workbook.xlsx")
        workbook.close()
    elif mutation == "chart":
        path = candidate / "workbook.xlsx"
        with zipfile.ZipFile(path) as archive:
            parts = {n: archive.read(n) for n in archive.namelist()}
        parts["xl/charts/chart1.xml"] = parts["xl/charts/chart1.xml"].replace(b"$B$3", b"$B$2")
        with zipfile.ZipFile(path, "w") as archive:
            for name, content in parts.items():
                archive.writestr(name, content)
    elif mutation == "asset":
        (candidate / "runtime.js").write_text("changed runtime")
    else:
        (candidate / "dashboard.html").write_text("<p>Missing plots</p>")
    assert not compare_artifacts(baseline, candidate)["equal"]


def test_artifact_manifest_requires_offline_assets_and_preserves_literal_formula(tmp_path):
    directory = tmp_path / "artifacts"
    _minimal_artifacts(directory)
    manifest = artifact_manifest(directory)
    assert manifest["sheets"][0]["rows"] == 3
    assert manifest["sheets"][0]["formulas"] == 0
    (directory / "runtime.js").unlink()
    with pytest.raises(ValueError, match="Missing or nonlocal"):
        artifact_manifest(directory)
    (directory / "dashboard.html").write_text('<script src="https://example.invalid/runtime.js"></script>')
    with pytest.raises(ValueError, match="Remote dashboard asset"):
        artifact_manifest(directory)


def test_artifact_metadata_parser_rejects_xml_entities():
    from defusedxml.common import DefusedXmlException

    payload = b'<!DOCTYPE x [<!ENTITY expansion "unexpected">]><x>&expansion;</x>'
    with pytest.raises(DefusedXmlException):
        _normalized_part("docProps/core.xml", payload)
