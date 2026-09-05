#!/usr/bin/env python3
"""Strict semantic artifact comparison for the synthetic CSV pipeline benchmark.

Only XLSX creation/modification timestamps and the ephemeral SQLite diagnostic
path are normalized. Worksheet data, formulas, styles, chart references, images,
HTML, scripts and offline assets must otherwise be identical.
"""
from __future__ import annotations

import argparse
import hashlib
from html.parser import HTMLParser
import json
from pathlib import Path
import re
import zipfile
from defusedxml import ElementTree as ET

_SQLITE_PATH = re.compile(
    rb"(?<='sqlite_path': ')/[^'<>\s]*/metroliza_csv_summary_[a-z0-9_]+\.sqlite(?=')"
)
_DCTERMS = "{http://purl.org/dc/terms/}"


def _normalized_part(name: str, payload: bytes) -> bytes:
    if name == "xl/sharedStrings.xml":
        return _SQLITE_PATH.sub(b"<temporary-sqlite>", payload)
    if name == "docProps/core.xml":
        root = ET.fromstring(payload)
        for tag in ("created", "modified"):
            element = root.find(_DCTERMS + tag)
            if element is not None:
                element.text = "<timestamp>"
        return ET.tostring(root)
    return payload


class _OfflineAssets(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.sources: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = dict(attrs)
        source = values.get("src") if tag in {"script", "img"} else None
        if source:
            self.sources.append(source)


def artifact_manifest(directory: Path) -> dict:
    """Read every XLSX part and referenced HTML asset, failing on incomplete output."""
    import openpyxl

    html_path = directory / "dashboard.html"
    html_bytes = html_path.read_bytes()
    parser = _OfflineAssets()
    parser.feed(html_bytes.decode("utf-8"))
    for source in parser.sources:
        if source.startswith("data:"):
            continue
        if "://" in source or source.startswith("//"):
            raise ValueError(f"Remote dashboard asset: {source}")
        target = (directory / source).resolve()
        if not target.is_relative_to(directory.resolve()) or not target.is_file():
            raise ValueError(f"Missing or nonlocal dashboard asset: {source}")

    workbook_path = directory / "workbook.xlsx"
    with zipfile.ZipFile(workbook_path) as archive:
        parts = {name: hashlib.sha256(_normalized_part(name, archive.read(name))).hexdigest()
                 for name in sorted(archive.namelist())}
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    sheets = []
    try:
        for sheet in workbook:
            digest = hashlib.sha256()
            formulas = 0
            nonempty = 0
            row_count = 0
            for row in sheet.iter_rows():
                row_count += 1
                cells = []
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        value = _SQLITE_PATH.sub(b"<temporary-sqlite>", value.encode()).decode()
                    nonempty += value is not None
                    formulas += cell.data_type == "f"
                    cells.append((cell.data_type, value))
                digest.update(json.dumps(cells, default=str, ensure_ascii=False).encode())
                digest.update(b"\n")
            sheets.append({"name": sheet.title, "rows": row_count, "columns": sheet.max_column,
                           "nonempty_cells": nonempty, "formulas": formulas,
                           "cells_sha256": digest.hexdigest()})
    finally:
        workbook.close()
    files = {str(path.relative_to(directory)): hashlib.sha256(path.read_bytes()).hexdigest()
             for path in sorted(directory.rglob("*"))
             if path.is_file() and path.suffix not in {".xlsx", ".pstats"}}
    return {"xlsx_parts": parts, "sheets": sheets, "offline_files": files,
            "referenced_assets": parser.sources}


def compare_artifacts(baseline: Path, candidate: Path) -> dict:
    before = artifact_manifest(baseline)
    after = artifact_manifest(candidate)
    differences = []
    for section in before:
        if before[section] != after[section]:
            differences.append(section)
    return {"equal": not differences, "differing_sections": differences,
            "baseline": before, "candidate": after}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("baseline", type=Path)
    parser.add_argument("candidate", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = compare_artifacts(args.baseline, args.candidate)
    args.output.write_text(json.dumps(result, indent=2) + "\n")
    print(json.dumps({"equal": result["equal"], "differing_sections": result["differing_sections"]}))
    raise SystemExit(0 if result["equal"] else 1)


if __name__ == "__main__":
    main()
